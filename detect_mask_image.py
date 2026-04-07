import cv2
import numpy as np
from tensorflow.keras.models import load_model
from tensorflow.keras.applications.mobilenet_v2 import preprocess_input
from tensorflow.keras.preprocessing.image import img_to_array
import sys
import os
import openpyxl
from openpyxl.styles import PatternFill, Font
from datetime import datetime

MODEL_PATH = "mask_detector.h5"
IMG_SIZE   = 224
CLASSES    = ["mask_weared_incorrect", "with_mask", "without_mask"]
COLORS     = {"with_mask": (0, 255, 0), "without_mask": (0, 0, 255), "mask_weared_incorrect": (0, 165, 255)}

print("[INFO] Loading model...")
import tensorflow as tf
model = tf.keras.models.load_model(MODEL_PATH, compile=False)
detector = cv2.CascadeClassifier(cv2.data.haarcascades + "haarcascade_frontalface_default.xml")

def detect_and_predict(image):
    gray  = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY)
    faces = detector.detectMultiScale(gray, scaleFactor=1.1, minNeighbors=5, minSize=(60, 60))

    results = []
    for (x, y, w, h) in faces:
        face = image[y:y+h, x:x+w]
        face = cv2.cvtColor(face, cv2.COLOR_BGR2RGB)
        face = cv2.resize(face, (IMG_SIZE, IMG_SIZE))
        face = img_to_array(face)
        face = preprocess_input(face)
        face = np.expand_dims(face, axis=0)

        preds = model.predict(face)[0]
        idx   = np.argmax(preds)
        label = CLASSES[idx]
        conf  = preds[idx]
        results.append(((x, y, w, h), label, conf))

    return results

# ── Main ─────────────────────────────────────────────────
img_path = sys.argv[1] if len(sys.argv) > 1 else "test.jpg"
if not os.path.exists(img_path):
    print(f"[ERROR] Image not found: {img_path}")
    sys.exit(1)

image   = cv2.imread(img_path)
results = detect_and_predict(image)

for ((x, y, w, h), label, conf) in results:
    color = COLORS.get(label, (255, 255, 255))
    text  = f"{label}: {conf*100:.1f}%"
    cv2.rectangle(image, (x, y), (x+w, y+h), color, 2)
    cv2.putText(image, text, (x, y - 10), cv2.FONT_HERSHEY_SIMPLEX, 0.55, color, 2)

# ── Save to Excel with two sides ──────────────────────────
excel_file = "mask_detection_log.xlsx"
now = datetime.now()

# Load or create workbook
if os.path.exists(excel_file):
    wb = openpyxl.load_workbook(excel_file)
    ws = wb.active
else:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Mask Detection Log"

    # ── Headers ──
    # Left side - Mask Worn
    ws["A1"] = "MASK WORN"
    ws["A1"].font = Font(bold=True, color="FFFFFF")
    ws["A1"].fill = PatternFill("solid", fgColor="008000")  # Green
    ws.merge_cells("A1:C1")

    ws["A2"] = "Image"
    ws["B2"] = "Confidence"
    ws["C2"] = "Date & Time"

    # Right side - No Mask
    ws["E1"] = "NO MASK / INCORRECT"
    ws["E1"].font = Font(bold=True, color="FFFFFF")
    ws["E1"].fill = PatternFill("solid", fgColor="FF0000")  # Red
    ws.merge_cells("E1:G1")

    ws["E2"] = "Image"
    ws["F2"] = "Confidence"
    ws["G2"] = "Date & Time"

    # Column widths
    for col in ["A", "B", "C", "E", "F", "G"]:
        ws.column_dimensions[col].width = 20

# Find next empty row for each side
mask_row = 3
no_mask_row = 3
for row in ws.iter_rows(min_row=3):
    if row[0].value:
        mask_row += 1
    if row[4].value:
        no_mask_row += 1

# ── Write data ────────────────────────────────────────────
for ((x, y, w, h), label, conf) in results:
    dt = now.strftime("%Y-%m-%d %H:%M:%S")
    if label == "with_mask":
        ws.cell(row=mask_row, column=1).value = img_path
        ws.cell(row=mask_row, column=2).value = f"{conf*100:.1f}%"
        ws.cell(row=mask_row, column=3).value = dt
        ws.cell(row=mask_row, column=1).fill = PatternFill("solid", fgColor="CCFFCC")
        mask_row += 1
    else:
        ws.cell(row=no_mask_row, column=5).value = img_path
        ws.cell(row=no_mask_row, column=6).value = f"{conf*100:.1f}%"
        ws.cell(row=no_mask_row, column=7).value = dt
        ws.cell(row=no_mask_row, column=5).fill = PatternFill("solid", fgColor="FFCCCC")
        no_mask_row += 1

wb.save(excel_file)
print(f"[INFO] Results saved to {excel_file}")

# ── Show image ────────────────────────────────────────────
image = cv2.resize(image, (350, 450))
cv2.namedWindow("Result", cv2.WINDOW_NORMAL)
cv2.resizeWindow("Result", 350, 450)
cv2.moveWindow("Result", 100, 100)
cv2.imshow("Result", image)
cv2.waitKey(0)
cv2.destroyAllWindows()
